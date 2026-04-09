"""
Maricopa County Motivated Seller Lead Scraper  v10.0

Exact workflow (confirmed from browser screenshots 2026-04-08):
  1. Search: recorder.maricopa.gov/recording/document-search-results.html
             ?documentCode=NS&beginDate=2026-04-08&endDate=2026-04-09
  2. Results page lists: Recording Number, Recording Date, Document Code
  3. PDF URL:  legacy.recorder.maricopa.gov/UnOfficialDocs/pdf/{recording_number}.pdf
  4. PDF contains: trustor name, address, APN, loan balance, beneficiary, sale date

Usage:
  python fetch.py              # full scrape → leads.xlsx + ghl_export.csv + JSON
  python fetch.py --test NOTS  # scrape NOTS only, print first 10 results
  python fetch.py --debug      # full scrape + save raw HTML/PDFs to data/debug/
  python fetch.py --diagnose   # probe page structure
"""

from __future__ import annotations
import argparse, asyncio, csv, io, json, logging, os, re, sys
import time, traceback, zipfile, tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import requests
from bs4 import BeautifulSoup

# ── Tesseract path — works on Windows and Linux (GitHub Actions) ─────────────
try:
    import pytesseract, os, sys
    if sys.platform == "win32":
        # Windows: hard-coded default install path
        _TESS = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.isfile(_TESS):
            pytesseract.pytesseract.tesseract_cmd = _TESS
    # Linux (GitHub Actions): tesseract is on PATH after apt-get install tesseract-ocr
    # pytesseract finds it automatically — no path needed
except ImportError:
    pass

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)])
log = logging.getLogger("maricopa")

# ── Config ────────────────────────────────────────────────────────────────────
LOOKBACK_DAYS = int(os.getenv("LOOKBACK_DAYS", "7"))

# New recorder site — search results URL template
RESULTS_URL = (
    "https://recorder.maricopa.gov/recording/document-search-results.html"
    "?lastNames=&firstNames=&middleNameIs=&documentTypeSelector=code"
    "&documentCode={code}&beginDate={start}&endDate={end}"
)

# PDF URL template (confirmed from screenshot)
PDF_URL = "https://legacy.recorder.maricopa.gov/UnOfficialDocs/pdf/{rec}.pdf"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

ASSESSOR_URLS = [
    "https://mcassessor.maricopa.gov/files/SFR_Parcel.zip",
    "https://mcassessor.maricopa.gov/files/Parcel.zip",
    "https://mcassessor.maricopa.gov/files/parcel_data.zip",
]
OUTPUT_PATHS  = [Path("dashboard/records.json"), Path("data/records.json")]
XLSX_PATH     = Path("data/leads.xlsx")
GHL_CSV_PATH  = Path("data/ghl_export.csv")
DEBUG_DIR     = Path("data/debug")

# ── Hard-coded doc codes (from --list-codes 2026-04-08) ───────────────────────
# Format: new-site code → (category, label)
CODES: dict[str, tuple[str, str]] = {
    "NS":  ("NOTS", "Notice of Trustee Sale"),
    "LP":  ("LP",   "Lis Pendens"),
    "LN":  ("LIEN", "Lien"),          # Governmental/Non-Govt Liens
    "FL":  ("LIEN", "Federal Tax Lien"),
    "SL":  ("LIEN", "State Tax Lien"),
    "ML":  ("LIEN", "Mechanic Lien"),
    "HL":  ("LIEN", "Medical Lien"),
    "JG":  ("JUD",  "Judgment"),
    "SJ":  ("JUD",  "Judgment"),       # Child Support Judgment
    "DE":  ("TAX",  "Tax Deed"),       # new site uses DE for Tax Deed
    "PD":  ("PRO",  "Probate Document"),
    "PJ":  ("PRO",  "Probate Document"),
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def _clean(s) -> str:
    return re.sub(r'\s+', ' ', str(s or "")).strip()

def _norm_date(raw: str) -> str:
    raw = _clean(raw)
    if not raw: return ""
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y",
                "%d/%m/%Y", "%B %d, %Y", "%b %d, %Y",
                "%-m-%d-%Y", "%m-%d-%Y"):
        try: return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError: pass
    # Handle "4-08-2026" style
    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', raw)
    if m:
        try: return datetime.strptime(f"{m.group(1)}/{m.group(2)}/{m.group(3)}", "%m/%d/%Y").strftime("%Y-%m-%d")
        except ValueError: pass
    return raw

def _parse_amount(raw) -> Optional[float]:
    clean = re.sub(r"[^\d.]", "", str(raw or ""))
    try:
        v = float(clean); return v if v > 0 else None
    except ValueError: return None

# ── Step 1: Fetch results page ────────────────────────────────────────────────

async def fetch_results(code: str, start_ymd: str, end_ymd: str,
                        save_debug: bool = False) -> list[dict]:
    """
    Load the search-results page via Playwright (needs JS to render),
    extract all recording numbers + dates.
    start_ymd / end_ymd must be YYYY-MM-DD.
    """
    from playwright.async_api import async_playwright

    url = RESULTS_URL.format(code=code, start=start_ymd, end=end_ymd)
    log.info(f"  [{code}] {url}")

    records: list[dict] = []

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"])
        ctx  = await browser.new_context(user_agent=UA)
        page = await ctx.new_page()

        try:
            await page.goto(url, timeout=60_000, wait_until="networkidle")
            await asyncio.sleep(2)

            page_num = 1
            while page_num <= 100:
                html = await page.content()

                if save_debug:
                    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
                    (DEBUG_DIR / f"{code}_results_p{page_num}.html").write_text(
                        html, encoding="utf-8")

                soup  = BeautifulSoup(html, "lxml")
                batch = _parse_results_page(soup, code)
                records.extend(batch)
                log.info(f"  [{code}] results page {page_num}: {len(batch)} records")

                if page_num == 1 and not batch:
                    # Log what the page says for debugging
                    body_text = " ".join(soup.get_text().split())[:400]
                    log.info(f"  [{code}] page text: {body_text}")

                # Next page button
                next_btn = None
                for sel in [
                    "a:has-text('Next')", "button:has-text('Next')",
                    "[aria-label='Next page']", ".pagination-next",
                    "a:has-text('›')", "a:has-text('»')",
                ]:
                    try:
                        el = await page.query_selector(sel)
                        if el and await el.is_visible():
                            # Check not disabled
                            cls = await el.get_attribute("class") or ""
                            if "disabled" not in cls.lower():
                                next_btn = sel; break
                    except Exception: pass

                if not next_btn: break
                await page.click(next_btn, timeout=5_000)
                await page.wait_for_load_state("networkidle", timeout=20_000)
                await asyncio.sleep(1)
                page_num += 1

        except Exception as exc:
            log.error(f"  [{code}] results fetch error: {exc}")
        finally:
            await browser.close()

    return records


def _parse_results_page(soup: BeautifulSoup, code: str) -> list[dict]:
    """Parse recording numbers and dates from the search results table."""
    records: list[dict] = []

    for table in soup.find_all("table"):
        headers = [th.get_text(" ", strip=True).lower()
                   for th in table.find_all("th")]
        if not any("recording" in h or "number" in h or "date" in h
                   for h in headers):
            continue

        for row in table.find_all("tr")[1:]:
            cells = row.find_all(["td", "th"])
            if len(cells) < 2: continue

            # Recording number — look for the link/button with the number
            rec_num = ""
            link    = ""
            for cell in cells:
                # Button or link with recording number
                for el in cell.find_all(["a", "button"]):
                    txt = _clean(el.get_text())
                    if re.match(r'\d{10,14}', txt):
                        rec_num = txt
                        href = el.get("href", "")
                        if href and href.startswith("http"):
                            link = href
                        break
                if rec_num: break

            if not rec_num:
                # Try first cell plain text
                txt = _clean(cells[0].get_text())
                if re.match(r'\d{10,14}', txt):
                    rec_num = txt

            if not rec_num: continue

            # Recording date — look for date pattern in cells
            filed = ""
            for cell in cells:
                txt = _clean(cell.get_text())
                # Match patterns like "4-08-2026" or "04/08/2026"
                if re.search(r'\d{1,2}[-/]\d{2}[-/]\d{4}', txt):
                    filed = _norm_date(txt)
                    break

            records.append({
                "doc_num":   rec_num,
                "doc_type":  code,
                "filed":     filed,
                "pdf_url":   PDF_URL.format(rec=rec_num),
                "clerk_url": link or f"https://recorder.maricopa.gov/recording/document-details.html?rec={rec_num}",
                "_code":     code,
            })

    # Also try direct div/list structure (React may not render a <table>)
    if not records:
        # Look for recording number patterns anywhere on page
        for el in soup.find_all(["div", "li", "tr", "span"],
                                string=re.compile(r'\d{11,13}')):
            txt = _clean(el.get_text())
            m = re.search(r'(\d{11,13})', txt)
            if m:
                rec_num = m.group(1)
                # Find date nearby
                parent_text = _clean(el.parent.get_text() if el.parent else "")
                date_m = re.search(r'(\d{1,2}[-/]\d{2}[-/]\d{4})', parent_text)
                filed = _norm_date(date_m.group(1)) if date_m else ""
                records.append({
                    "doc_num":   rec_num,
                    "doc_type":  code,
                    "filed":     filed,
                    "pdf_url":   PDF_URL.format(rec=rec_num),
                    "clerk_url": f"https://recorder.maricopa.gov/recording/document-details.html?rec={rec_num}",
                    "_code":     code,
                })

    return records

# ── Step 2: Parse PDF ─────────────────────────────────────────────────────────

class PDFParser:
    """Download and parse a NOTS PDF from legacy.recorder.maricopa.gov"""

    # Patterns to extract from NOTS PDFs (confirmed from real OCR output 2026-04-08)
    _PAT = {
        # Structure confirmed from OCR:
        # "Name and address of original trustor:\n(as shown on the Deed of Trust)\nBRANDON S. MAXWELL..."
        # The [^\n]* after the label matches the rest of that line (empty or colon)
        # Then \([^\n]+\)\n skips the parenthetical line
        "trustor": [
            re.compile(
                r'Name\s+and\s+address\s+of\s+original\s+trustor[^\n]*\n'
                r'\([^\n]+\)\n'                 # skip "(as shown on the Deed of Trust)"
                r'([A-Z][^\n]{3,120})',
                re.I | re.MULTILINE),
            re.compile(
                r'Name\s+and\s+address\s+of\s+original\s+trustor[^\n]*\n'
                r'([A-Z][^\n]{3,120})',
                re.I | re.MULTILINE),
        ],
        # "Name and address of beneficiary:\n(as of recording of Notice of Sale)\nFreedom Mortgage..."
        "beneficiary": [
            re.compile(
                r'Name\s+and\s+address\s+of\s+beneficiar[^\n]*\n'
                r'\([^\n]+\)\n'                 # skip "(as of recording...)"
                r'([^\n]{3,120})',
                re.I | re.MULTILINE),
            re.compile(
                r'Name\s+and\s+address\s+of\s+beneficiar[^\n]*\n'
                r'([^\n]{3,120})',
                re.I | re.MULTILINE),
        ],
        # "Original Principal Balance: $418,396.00"
        "amount": [
            re.compile(
                r'Original\s+Principal\s+Balance[:\s]+\$?([\d,]+(?:\.\d{1,2})?)',
                re.I),
            re.compile(
                r'(?:Unpaid\s+(?:Principal\s+)?Balance|Total\s+(?:Unpaid\s+)?'
                r'(?:Debt|Indebtedness))[:\s]+\$?([\d,]+(?:\.\d{1,2})?)',
                re.I),
        ],
        # "on 7/8/2026 at 12:00 PM"
        "sale_date": [
            re.compile(r'\bon\s+(\d{1,2}/\d{1,2}/\d{4})\s+at\s+\d{1,2}:\d{2}', re.I),
            re.compile(r'\bon\s+(\w+\s+\d{1,2},?\s*\d{4})\s+at\s+\d{1,2}:\d{2}', re.I),
        ],
        # "Superior Court Building, 201 W. Jefferson, Phoenix, AZ 85003"
        "sale_location": [
            re.compile(
                r'((?:Superior\s+Court|Maricopa\s+County)[^\n,]{0,60}'
                r'(?:Building|Courthouse|Jefferson)[^\n]{0,80})',
                re.I),
            re.compile(
                r'(?:courtyard|main\s+entrance|highest\s+bidder)[^\n]{0,60}\n?'
                r'([^\n]{10,120}(?:Jefferson|Court|Building)[^\n]{0,60})',
                re.I),
        ],
        # APN — present in some NOTS PDFs
        "apn": [
            re.compile(r'Tax\s+Parcel\s+#?\s*[:\s]+([0-9\-]{5,15})', re.I),
            re.compile(
                r'(?:APN|Parcel\s+(?:No\.?|Number))[:\s#]+(\d{3}[\-\s]?\d{2}[\-\s]?\d{3,4})',
                re.I),
        ],
        # Trustee
        "trustee": [
            re.compile(
                r'NAME,\s*ADDRESS[^:\n]+TRUSTEE[^\n]*\n'
                r'(?:\([^\n]+\)\n)?'
                r'([^\n]{3,80})',
                re.I),
        ],
    }

    def __init__(self, session: requests.Session):
        self._s = session

    def parse(self, pdf_url: str, rec_num: str = "") -> dict:
        out: dict = {k: "" for k in self._PAT}
        out.update({"_pdf_ok": False, "_ocr_used": False,
                    "prop_city": "", "prop_state": "AZ", "prop_zip": ""})

        try:
            pdf_bytes = self._download(pdf_url)
            if not pdf_bytes:
                return out

            text = self._extract_text(pdf_bytes)
            if len(text.strip()) < 100:
                text = self._ocr(pdf_bytes)
                out["_ocr_used"] = True

            if len(text.strip()) > 50:
                out["_pdf_ok"] = True
                self._apply_patterns(text, out)
                self._post_process(out, text)

        except Exception as exc:
            log.debug(f"PDF error ({pdf_url}): {exc}")

        return out

    def _download(self, url: str) -> Optional[bytes]:
        for attempt in range(3):
            try:
                r = self._s.get(url, timeout=30, stream=True)
                if r.status_code == 200:
                    content = r.content
                    if content[:4] == b'%PDF':
                        return content
                    else:
                        log.debug(f"PDF not a PDF ({url}): {content[:100]}")
                        return None
                log.debug(f"PDF HTTP {r.status_code}: {url}")
            except Exception as exc:
                log.debug(f"PDF download attempt {attempt+1}: {exc}")
                time.sleep(1)
        return None

    @staticmethod
    def _extract_text(pdf_bytes: bytes) -> str:
        try:
            import pdfplumber
            parts = []
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for p in pdf.pages[:6]:
                    t = p.extract_text(x_tolerance=2, y_tolerance=2) or ""
                    parts.append(t)
            text = "\n".join(parts)
            if len(text.strip()) > 50:
                return text
        except Exception: pass
        try:
            from pdfminer.high_level import extract_text
            return extract_text(io.BytesIO(pdf_bytes)) or ""
        except Exception: return ""

    @staticmethod
    def _ocr(pdf_bytes: bytes) -> str:
        try:
            import fitz, pytesseract
            from PIL import Image
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            parts = []
            for i in range(min(len(doc), 4)):
                pix = doc[i].get_pixmap(matrix=fitz.Matrix(2.5, 2.5))
                img = Image.open(io.BytesIO(pix.tobytes("png")))
                parts.append(pytesseract.image_to_string(img, config="--oem 3 --psm 6") or "")
            doc.close()
            return "\n".join(parts)
        except Exception: return ""

    def _apply_patterns(self, text: str, out: dict):
        for field, patterns in self._PAT.items():
            for pat in patterns:
                m = pat.search(text)
                if m:
                    val = _clean(m.group(1) if m.lastindex else m.group())
                    if val and len(val) > 2:
                        out[field] = val
                        break

    @staticmethod
    def _post_process(out: dict, full_text: str):
        # Clean trustor — name may span multiple lines; strip everything after the address
        if out["trustor"]:
            # Remove everything from ", HUSBAND AND WIFE" or similar legal boilerplate onward
            out["trustor"] = re.sub(
                r',?\s+(?:HUSBAND\s+AND\s+WIFE|A\s+MARRIED|A\s+SINGLE|AS\s+TRUSTEE'
                r'|COMMUNITY\s+PROPERTY|NOT\s+AS|AND\s+NOT).*',
                "", out["trustor"], flags=re.I).strip()

        # Extract property address from the trustor block
        # Structure after name lines: STREET\nCITY, AZ ZIP
        if not out.get("property_address") and out.get("trustor"):
            m = re.search(
                r'Name\s+and\s+address\s+of\s+original\s+trustor[^\n]*\n'
                r'(?:\([^\n]+\)\n)?'           # skip parenthetical
                r'(?:[A-Z][^\n]+\n){1,5}?'     # skip 1-5 name lines
                r'(\d+\s+[^\n]{5,60})\n'       # street: starts with number
                r'([A-Z][A-Za-z\s]+),\s*(AZ|Arizona)\s+(\d{5})',
                full_text, re.I | re.MULTILINE)
            if m:
                out["property_address"] = _clean(m.group(1))
                out["prop_city"]        = _clean(m.group(2))
                out["prop_state"]       = "AZ"
                out["prop_zip"]         = m.group(4)

        # Clean trustee — remove junk that OCR picks up after the name
        if out.get("trustee"):
            out["trustee"] = re.sub(
                r'\s*(?:Unofficial\s+Document|Sale\s+(?:Lines|Information)|'
                r'as\s+of\s+recording|SALE\s+INFORMATION).*',
                "", out["trustee"], flags=re.I).strip().rstrip(",")

        # Truncate sale_location to just the building address (before ", on")
        if out.get("sale_location"):
            loc = re.sub(r',?\s+on\s+\d{1,2}/\d{1,2}/\d{4}.*', "",
                         out["sale_location"], flags=re.I).strip()
            out["sale_location"] = loc[:120]
        if out.get("apn"):
            d = re.sub(r'\D', '', out["apn"])
            if len(d) >= 8:
                out["apn"] = f"{d[:3]}-{d[3:5]}-{d[5:]}"

        # Normalise sale date
        if out.get("sale_date"):
            out["sale_date"] = _norm_date(out["sale_date"])


# ── Parcel Index ──────────────────────────────────────────────────────────────

class ParcelIndex:
    def __init__(self):
        self._by_owner: dict[str, list[dict]] = {}
        self._by_apn:   dict[str, dict]       = {}

    def load(self) -> bool:
        for url in ASSESSOR_URLS:
            try:
                log.info(f"Downloading parcel data: {url}")
                r = requests.get(url, timeout=120, stream=True,
                                 headers={"User-Agent": UA})
                if r.status_code != 200:
                    log.warning(f"  HTTP {r.status_code}: {url}"); continue
                log.info(f"  Downloaded {len(r.content):,} bytes")
                if self._parse_zip(r.content): return True
            except Exception as exc:
                log.warning(f"  Failed ({url}): {exc}")
        log.warning("Parcel data unavailable — address enrichment disabled")
        return False

    def _parse_zip(self, raw: bytes) -> bool:
        try: zf = zipfile.ZipFile(io.BytesIO(raw))
        except Exception: return False
        names = zf.namelist()
        dbfs = [n for n in names if n.lower().endswith(".dbf")]
        csvs = [n for n in names if n.lower().endswith(".csv")]
        if dbfs: return self._parse_dbf(zf, dbfs[0])
        if csvs: return self._parse_csv(zf, csvs[0])
        return False

    def _parse_dbf(self, zf, name: str) -> bool:
        try:
            from dbfread import DBF
        except ImportError:
            log.warning("dbfread not installed"); return False
        raw = zf.read(name)
        tmp = Path(tempfile.mktemp(suffix=".dbf"))
        tmp.write_bytes(raw); count = 0
        try:
            for rec in DBF(str(tmp), encoding="latin-1",
                           ignore_missing_memofile=True):
                self._ingest(dict(rec)); count += 1
        except Exception as exc:
            log.error(f"DBF error: {exc}")
        finally:
            tmp.unlink(missing_ok=True)
        log.info(f"  Loaded {count:,} parcel records (DBF)"); return count > 0

    def _parse_csv(self, zf, name: str) -> bool:
        raw = zf.read(name).decode("latin-1", errors="replace"); count = 0
        for row in csv.DictReader(io.StringIO(raw)):
            try: self._ingest(row); count += 1
            except Exception: pass
        log.info(f"  Loaded {count:,} parcel records (CSV)"); return count > 0

    def _ingest(self, rec: dict):
        def g(*keys):
            for k in keys:
                for var in (k, k.upper(), k.lower()):
                    v = rec.get(var, "")
                    if v and str(v).strip() not in ("", "None"):
                        return str(v).strip()
            return ""
        apn   = g("APN","PARCELNO","PARCEL_NO","PARCEL")
        owner = g("OWNER","OWN1","OWNNAME","OWNER_NAME")
        p = {
            "apn":        self._norm(apn), "owner": owner,
            "site_addr":  g("SITE_ADDR","SITEADDR","SITE_ADDRESS","SITUS_ADDR"),
            "site_city":  g("SITE_CITY","SITECITY"),
            "site_state": "AZ",
            "site_zip":   g("SITE_ZIP","SITEZIP"),
            "mail_addr":  g("ADDR_1","MAILADR1","MAIL_ADDR"),
            "mail_city":  g("CITY","MAILCITY","MAIL_CITY"),
            "mail_state": g("STATE","MAILSTATE") or "AZ",
            "mail_zip":   g("ZIP","MAILZIP","MAIL_ZIP"),
        }
        if self._norm(apn): self._by_apn[self._norm(apn)] = p
        if owner:
            for v in self._variants(owner):
                self._by_owner.setdefault(v, []).append(p)

    @staticmethod
    def _norm(apn: str) -> str:
        return re.sub(r'\D', '', apn or "")

    @staticmethod
    def _variants(name: str) -> list[str]:
        name = name.strip().upper(); v = {name}
        if "," in name:
            parts = [p.strip() for p in name.split(",", 1)]
            last, first = parts[0], parts[1]
            v.update({f"{first} {last}", f"{last} {first}", last})
        else:
            toks = name.split()
            if len(toks) >= 2:
                v.update({f"{toks[-1]} {' '.join(toks[:-1])}",
                           f"{toks[-1]}, {' '.join(toks[:-1])}"})
        return list(v)

    def lookup(self, name: str) -> Optional[dict]:
        if not name: return None
        upper = name.strip().upper()
        for v in self._variants(upper):
            hits = self._by_owner.get(v)
            if hits: return hits[0]
        for tok in [t for t in upper.split() if len(t) > 3][:2]:
            for key, parcels in self._by_owner.items():
                if tok in key: return parcels[0]
        return None

    def lookup_apn(self, apn: str) -> Optional[dict]:
        n = self._norm(apn)
        return self._by_apn.get(n) if n else None


# ── Scoring ───────────────────────────────────────────────────────────────────

def _apply_parcel(rec: dict, parcel: Optional[dict], pdf_addr: str = "",
                  pdf_city: str = "", pdf_zip: str = ""):
    if parcel:
        rec.setdefault("prop_address", parcel.get("site_addr", ""))
        rec.setdefault("prop_city",    parcel.get("site_city", ""))
        rec.setdefault("prop_state",   "AZ")
        rec.setdefault("prop_zip",     parcel.get("site_zip", ""))
        rec.setdefault("mail_address", parcel.get("mail_addr", ""))
        rec.setdefault("mail_city",    parcel.get("mail_city", ""))
        rec.setdefault("mail_state",   parcel.get("mail_state", "AZ"))
        rec.setdefault("mail_zip",     parcel.get("mail_zip", ""))
    # Use PDF address as fallback
    if pdf_addr and not rec.get("prop_address"):
        rec["prop_address"] = pdf_addr
        if pdf_city: rec["prop_city"] = pdf_city
        if pdf_zip:  rec["prop_zip"]  = pdf_zip
    # Defaults
    for k, d in [("prop_address",""),("prop_city",""),("prop_state","AZ"),
                 ("prop_zip",""),("mail_address",""),("mail_city",""),
                 ("mail_state","AZ"),("mail_zip","")]:
        rec.setdefault(k, d)


def score_record(rec: dict, all_records: list[dict]) -> tuple[list[str], int]:
    flags: list[str] = []; score = 30
    cat   = rec.get("cat", "")
    owner = (rec.get("owner") or "").upper()
    amt   = rec.get("amount") or 0
    filed = rec.get("filed", "")

    if cat == "LP":   flags.append("Lis pendens");      score += 10
    if cat in ("FC","NOTS"): flags.append("Pre-foreclosure"); score += 10
    if cat == "NOTS": flags.append("Trustee Sale");     score += 5
    if cat == "TAX":  flags.append("Tax deed");         score += 10
    if cat == "JUD":  flags.append("Judgment lien");    score += 10
    if cat == "LIEN":
        lbl = rec.get("cat_label","")
        if "Federal" in lbl or "State" in lbl:
            flags.append("Gov/Tax lien"); score += 10
        elif "Mechanic" in lbl: flags.append("Mechanic lien"); score += 8
        elif "Medical"  in lbl: flags.append("Medical lien");  score += 7
        else:                   flags.append("Lien");           score += 6
    if cat == "PRO":  flags.append("Probate / estate"); score += 10

    if cat in ("LP","FC","NOTS") and owner:
        owner_cats = {r.get("cat") for r in all_records
                      if r.get("owner","").upper() == owner
                      and r.get("doc_num") != rec.get("doc_num")}
        if ("LP" in owner_cats and cat in ("FC","NOTS")) or \
           (cat == "LP" and owner_cats & {"FC","NOTS"}): score += 20

    if amt > 100_000:  score += 15
    elif amt > 50_000: score += 10

    try:
        if (datetime.now() - datetime.strptime(filed, "%Y-%m-%d")).days <= 7:
            flags.append("New this week"); score += 5
    except Exception: pass

    if rec.get("prop_address") or rec.get("mail_address"): score += 5
    if rec.get("_pdf_ok"): score += 3
    if owner and any(kw in owner for kw in
                     ("LLC","INC","CORP"," LP","LTD","TRUST","HOLDINGS","PROPERTIES")):
        flags.append("LLC / corp owner")

    return flags, min(score, 100)


# ── Excel Export ──────────────────────────────────────────────────────────────

def write_xlsx(records: list[dict], path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed"); return

    COLS = [
        ("Score",             "score"),
        ("Category",          "cat_label"),
        ("Doc Type",          "doc_type"),
        ("Recording Number",  "doc_num"),
        ("Date Filed",        "filed"),
        ("Trustor / Owner",   "owner"),
        ("Beneficiary",       "grantee"),
        ("Loan Balance ($)",  "amount"),
        ("Sale Date",         "sale_date"),
        ("Sale Location",     "sale_location"),
        ("Property Address",  "prop_address"),
        ("Property City",     "prop_city"),
        ("Property State",    "prop_state"),
        ("Property Zip",      "prop_zip"),
        ("Mailing Address",   "mail_address"),
        ("Mailing City",      "mail_city"),
        ("Mailing State",     "mail_state"),
        ("Mailing Zip",       "mail_zip"),
        ("APN",               "apn"),
        ("Trustee",           "trustee"),
        ("Flags",             "_flags"),
        ("PDF Parsed",        "_pok"),
        ("PDF / Doc URL",     "clerk_url"),
    ]
    WIDTHS = {
        "Score":9,"Category":22,"Doc Type":10,"Recording Number":16,
        "Date Filed":12,"Trustor / Owner":34,"Beneficiary":30,
        "Loan Balance ($)":16,"Sale Date":13,"Sale Location":32,
        "Property Address":34,"Property City":18,"Property State":8,
        "Property Zip":9,"Mailing Address":34,"Mailing City":18,
        "Mailing State":8,"Mailing Zip":9,"APN":14,"Trustee":28,
        "Flags":40,"PDF Parsed":9,"PDF / Doc URL":12,
    }

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    HDR_FILL = PatternFill("solid", fgColor="1E2433")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HDR_ALGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"))
    DF = Font(name="Arial", size=9)
    UF = Font(name="Arial", size=9, color="2563EB", underline="single")
    CA = Alignment(horizontal="center", vertical="top")
    LA = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    ws = wb.active; ws.title = "Leads"
    ws.append([c[0] for c in COLS])
    for cell in ws[1]:
        cell.font=HDR_FONT; cell.fill=HDR_FILL
        cell.alignment=HDR_ALGN; cell.border=THIN
    ws.row_dimensions[1].height = 36

    for ri, rec in enumerate(records, 2):
        sc   = rec.get("score", 0)
        fill = (PatternFill("solid", fgColor="FEE2E2") if sc >= 70 else
                PatternFill("solid", fgColor="FEF9C3") if sc >= 50 else
                PatternFill("solid", fgColor="FFFFFF"))
        row = []
        for _, k in COLS:
            if k == "_flags": row.append("; ".join(rec.get("flags", [])))
            elif k == "_pok": row.append("Yes" if rec.get("_pdf_ok") else "No")
            else:             row.append(rec.get(k, "") or "")
        ws.append(row)
        for ci, cell in enumerate(ws[ri], 1):
            hdr = COLS[ci-1][0]; cell.border = THIN; cell.fill = fill
            if hdr == "Score":
                cell.font = Font(name="Arial", size=10, bold=True,
                    color=("C0392B" if sc>=70 else "B45309" if sc>=50 else "374151"))
                cell.alignment = CA
            elif hdr == "PDF / Doc URL" and cell.value:
                cell.hyperlink = str(cell.value); cell.value = "View ↗"
                cell.font = UF; cell.alignment = CA
            elif hdr == "Loan Balance ($)" and cell.value:
                try:
                    cell.value = float(str(cell.value).replace(",", ""))
                    cell.number_format = '"$"#,##0.00'
                except Exception: pass
                cell.font = Font(name="Arial", size=9, bold=True)
                cell.alignment = CA
            else:
                cell.font = DF; cell.alignment = LA
        ws.row_dimensions[ri].height = 26

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    for ci, (h, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(h, 14)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Maricopa County Motivated Seller Leads"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True)
    ws2["A2"] = f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    rows = [
        ("", ""), ("Total Leads", len(records)),
        ("With Property Address", sum(1 for r in records if r.get("prop_address"))),
        ("With Mailing Address",  sum(1 for r in records if r.get("mail_address"))),
        ("PDF Parsed",            sum(1 for r in records if r.get("_pdf_ok"))),
        ("Score 70+ (Hot)",       sum(1 for r in records if r.get("score",0) >= 70)),
        ("Score 50-69 (Warm)",    sum(1 for r in records if 50 <= r.get("score",0) < 70)),
        ("", ""), ("By Category", "Count"),
    ]
    cats: dict[str, int] = {}
    for r in records:
        l = r.get("cat_label", "Other"); cats[l] = cats.get(l, 0) + 1
    for l, c in sorted(cats.items(), key=lambda x: -x[1]):
        rows.append((l, c))
    for i, (a, b) in enumerate(rows, 3):
        ws2.cell(i, 1, a).font = Font(name="Arial", size=10, bold=bool(a))
        ws2.cell(i, 2, b).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12
    wb.save(path)
    log.info(f"Excel → {path} ({len(records)} rows)")


# ── GHL CSV ───────────────────────────────────────────────────────────────────

def write_ghl_csv(records: list[dict], path: Path):
    COLS = [
        "First Name","Last Name","Mailing Address","Mailing City","Mailing State",
        "Mailing Zip","Property Address","Property City","Property State","Property Zip",
        "Lead Type","Document Type","Date Filed","Recording Number","Loan Balance",
        "Sale Date","APN","Seller Score","Motivated Seller Flags","Source","PDF URL",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS); w.writeheader()
        for r in records:
            owner = r.get("owner") or ""
            if "," in owner:
                parts = owner.split(",", 1)
                fn, ln = parts[1].strip().title(), parts[0].strip().title()
            else:
                toks = owner.strip().split()
                fn = toks[0].title() if toks else ""
                ln = " ".join(toks[1:]).title() if len(toks) > 1 else ""
            amt = r.get("amount")
            w.writerow({
                "First Name":           fn,
                "Last Name":            ln,
                "Mailing Address":      r.get("mail_address", ""),
                "Mailing City":         r.get("mail_city", ""),
                "Mailing State":        r.get("mail_state", ""),
                "Mailing Zip":          r.get("mail_zip", ""),
                "Property Address":     r.get("prop_address", ""),
                "Property City":        r.get("prop_city", ""),
                "Property State":       r.get("prop_state", "AZ"),
                "Property Zip":         r.get("prop_zip", ""),
                "Lead Type":            r.get("cat_label", ""),
                "Document Type":        r.get("doc_type", ""),
                "Date Filed":           r.get("filed", ""),
                "Recording Number":     r.get("doc_num", ""),
                "Loan Balance":         f"{amt:,.2f}" if amt else "",
                "Sale Date":            r.get("sale_date", ""),
                "APN":                  r.get("apn", ""),
                "Seller Score":         r.get("score", ""),
                "Motivated Seller Flags": "; ".join(r.get("flags", [])),
                "Source":               "Maricopa County Recorder",
                "PDF URL":              r.get("pdf_url", r.get("clerk_url", "")),
            })
    log.info(f"GHL CSV → {path} ({len(records)} rows)")


# ── Diagnose ──────────────────────────────────────────────────────────────────

async def run_diagnose():
    from playwright.async_api import async_playwright
    # Load yesterday's NOTS results page directly
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    today     = datetime.now().strftime("%Y-%m-%d")
    url = RESULTS_URL.format(code="NS", start=yesterday, end=today)
    print(f"\nLoading: {url}")
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True,
            args=["--no-sandbox","--disable-dev-shm-usage","--disable-gpu"])
        ctx  = await browser.new_context(user_agent=UA)
        page = await ctx.new_page()
        try:
            await page.goto(url, timeout=60_000, wait_until="networkidle")
            await asyncio.sleep(2)
            print(f"Title: {await page.title()}")
            html = await page.content()
            soup = BeautifulSoup(html, "lxml")
            print(f"Page text (first 2000): {' '.join(soup.get_text().split())[:2000]}")
            print(f"\nTables: {len(soup.find_all('table'))}")
            for t in soup.find_all("table"):
                print(f"  headers: {[th.get_text() for th in t.find_all('th')][:8]}")
        except Exception as exc:
            print(f"Error: {exc}")
        finally:
            await browser.close()


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(description="Maricopa Motivated Seller Scraper v10")
    parser.add_argument("--test",     metavar="CAT",
        help="Scrape one category (NOTS, LP, JUD, LIEN, TAX, PRO) and print results")
    parser.add_argument("--debug",    action="store_true",
        help="Full scrape + save raw HTML/PDFs to data/debug/")
    parser.add_argument("--diagnose", action="store_true",
        help="Probe page structure of results page")
    args = parser.parse_args()

    if args.diagnose:
        await run_diagnose(); return

    # ── Check PDF parsing dependencies ────────────────────────────────────────
    pdf_lib = None
    try:
        import pdfplumber; pdf_lib = "pdfplumber"
    except ImportError:
        try:
            from pdfminer.high_level import extract_text; pdf_lib = "pdfminer"
        except ImportError:
            pass

    if not pdf_lib:
        log.error("="*60)
        log.error("PDF PARSING NOT AVAILABLE")
        log.error("Install pdfplumber:  pip install pdfplumber")
        log.error("This is required to extract owner/address from NOTS PDFs")
        log.error("="*60)
    else:
        log.info(f"PDF library: {pdf_lib} ✓")

    end_dt    = datetime.now()
    start_dt  = end_dt - timedelta(days=LOOKBACK_DAYS)
    # New site needs YYYY-MM-DD dates
    start_ymd = start_dt.strftime("%Y-%m-%d")
    end_ymd   = end_dt.strftime("%Y-%m-%d")

    log.info("=" * 60)
    log.info("Maricopa Motivated Seller Scraper v10.0")
    log.info(f"Date range: {start_ymd} → {end_ymd}")
    log.info("=" * 60)

    # Filter codes for --test
    codes_to_run = dict(CODES)
    if args.test:
        cat_filter = args.test.upper()
        codes_to_run = {k: v for k, v in CODES.items()
                        if v[0] == cat_filter}
        if not codes_to_run:
            log.error(f"No codes for '{args.test}'. Valid: "
                      f"{sorted(set(v[0] for v in CODES.values()))}")
            return

    # Step 1: collect recording numbers from results pages
    all_raw: list[dict] = []
    for code, (cat, cat_label) in codes_to_run.items():
        try:
            recs = await fetch_results(code, start_ymd, end_ymd,
                                       save_debug=args.debug)
            for r in recs:
                r["cat"]       = cat
                r["cat_label"] = cat_label
            all_raw.extend(recs)
            log.info(f"  ✓ {code} ({cat_label}): {len(recs)} recording numbers")
        except Exception as exc:
            log.error(f"  ✗ {code}: {exc}")

    log.info(f"Total recording numbers collected: {len(all_raw)}")

    # Create session early so we can test PDFs
    session = requests.Session()
    session.headers.update({
        "User-Agent": UA,
        "Referer":    "https://recorder.maricopa.gov/",
        "Accept":     "application/pdf,*/*",
    })
    try:
        session.get("https://recorder.maricopa.gov/recording/document-search.html",
                    timeout=10)
    except Exception:
        pass

    # Quick PDF sanity check on first record
    if all_raw:
        test_url = all_raw[0].get("pdf_url", "")
        if test_url:
            log.info(f"Testing PDF download: {test_url}")
            try:
                test_r = session.get(test_url, timeout=15)
                if test_r.status_code == 200 and test_r.content[:4] == b'%PDF':
                    log.info(f"  ✓ PDF accessible ({len(test_r.content):,} bytes)")
                else:
                    log.warning(f"  ✗ PDF failed — HTTP {test_r.status_code}, "
                                f"content start: {test_r.content[:120]}")
            except Exception as exc:
                log.warning(f"  ✗ PDF test error: {exc}")

    # Dedup
    seen: set[str] = set(); deduped: list[dict] = []
    for r in all_raw:
        key = r.get("doc_num", "")
        if key and key in seen: continue
        if key: seen.add(key)
        deduped.append(r)
    log.info(f"After dedup: {len(deduped)}")

    # Step 2: parse PDFs to get all fields
    pdf_parser = PDFParser(session)
    parcels    = ParcelIndex()
    parcel_ok  = parcels.load()

    enriched: list[dict] = []
    pdf_count = 0

    for i, rec in enumerate(deduped):
        pdf_url = rec.get("pdf_url", "")
        if pdf_url:
            try:
                pdf = pdf_parser.parse(pdf_url, rec.get("doc_num",""))
                # Merge PDF data into record
                if pdf.get("trustor"):
                    rec["owner"]   = pdf["trustor"]
                if pdf.get("beneficiary"):
                    rec["grantee"] = pdf["beneficiary"]
                if pdf.get("amount"):
                    v = _parse_amount(pdf["amount"])
                    if v: rec["amount"] = v
                if pdf.get("sale_date"):
                    rec["sale_date"] = pdf["sale_date"]
                if pdf.get("sale_location"):
                    rec["sale_location"] = pdf["sale_location"]
                if pdf.get("trustee"):
                    rec["trustee"] = pdf["trustee"]
                rec["apn"]      = pdf.get("apn","")
                rec["_pdf_ok"]  = pdf.get("_pdf_ok", False)

                # Parcel lookup: APN first, then name
                parcel = (parcels.lookup_apn(pdf.get("apn","")) or
                          parcels.lookup(rec.get("owner","")))
                _apply_parcel(rec, parcel,
                              pdf.get("property_address",""),
                              pdf.get("prop_city",""),
                              pdf.get("prop_zip",""))
                pdf_count += 1
            except Exception as exc:
                log.debug(f"PDF error ({pdf_url}): {exc}")
                _apply_parcel(rec, parcels.lookup(rec.get("owner","")))
        else:
            _apply_parcel(rec, parcels.lookup(rec.get("owner","")))

        if (i+1) % 20 == 0:
            log.info(f"  Processed {i+1}/{len(deduped)} PDFs...")

        enriched.append(rec)

    log.info(f"PDFs parsed: {pdf_count}/{len(deduped)}")

    # Score
    for rec in enriched:
        try:
            flags, score = score_record(rec, enriched)
            rec["flags"] = flags; rec["score"] = score
        except Exception:
            rec.setdefault("flags",[]); rec.setdefault("score",30)
    enriched.sort(key=lambda r: r.get("score",0), reverse=True)

    if args.test:
        print(f"\n=== TEST '{args.test}': {len(enriched)} records ===")
        for r in enriched[:10]:
            print(json.dumps(
                {k:v for k,v in r.items() if not k.startswith("__")},
                indent=2, default=str))
        return

    # Save
    with_addr = sum(1 for r in enriched
                    if r.get("prop_address") or r.get("mail_address"))
    payload = {
        "fetched_at":   datetime.utcnow().isoformat()+"Z",
        "source":       "Maricopa County Recorder",
        "date_range":   {"start":start_ymd,"end":end_ymd},
        "total":        len(enriched),
        "with_address": with_addr,
        "records":      enriched,
    }
    for path in OUTPUT_PATHS:
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(json.dumps(payload, indent=2, default=str))
            log.info(f"JSON → {path}")
        except Exception as exc: log.error(f"JSON save failed: {exc}")

    try:    write_xlsx(enriched, XLSX_PATH)
    except Exception as exc: log.error(f"Excel failed: {exc}")
    try:    write_ghl_csv(enriched, GHL_CSV_PATH)
    except Exception as exc: log.error(f"GHL CSV failed: {exc}")

    log.info(f"DONE — {len(enriched)} leads | {with_addr} with address | "
             f"parcel: {'✓' if parcel_ok else '✗'} | PDFs: {pdf_count}")


if __name__ == "__main__":
    asyncio.run(main())
